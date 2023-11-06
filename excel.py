
import cv2
from openpyxl import *

from listar import *
import os

from tamanho import *
from excel import *
import os.path


def comparaind(nome, pontos, razao,espaco):
    img = cv2.imread(nome)
    if (espaco == "hsv"):
        img = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

    elif (espaco == "hls"):
        img = cv2.cvtColor(img, cv2.COLOR_BGR2HLS)




    # roi1 pontoteste[0][0]
    roi1 = img[pontos[0][0][1] - (5 * razao):pontos[0][0][1] + (5 * razao),
           pontos[0][0][0] - (5 * razao):pontos[0][0][0] + (5 * razao)]
    roi2 = img[pontos[0][1][1] - (5 * razao):pontos[0][1][1] + (5 * razao),
           pontos[0][1][0] - (5 * razao):pontos[0][1][0] + (5 * razao)]
    roi3 = img[pontos[0][2][1] - (5 * razao):pontos[0][2][1] + (5 * razao),
           pontos[0][2][0] - (5 * razao):pontos[0][2][0] + (5 * razao)]

    # distante da area ponto teste[1]
    roi4 = img[pontos[1][0][1] - (5 * razao):pontos[1][0][1] + (5 * razao),
           pontos[1][0][0] - (5 * razao):pontos[1][0][0] + (5 * razao)]
    roi5 = img[pontos[1][1][1] - (5 * razao):pontos[1][1][1] + (5 * razao),
           pontos[1][1][0] - (5 * razao):pontos[1][1][0] + (5 * razao)]
    roi6 = img[pontos[1][2][1] - (5 * razao):pontos[1][2][1] + (5 * razao),
           pontos[1][2][0] - (5 * razao):pontos[1][2][0] + (5 * razao)]

    """if(espaco=='rgb'):
        cv2.imwrite("fotos/detalhado/roi1"+espaco+".png" , roi1)
        cv2.imwrite( "fotos/detalhado/roi2"+espaco+".png" , roi2)
        cv2.imwrite( "fotos/detalhado/roi3"+espaco+".png" , roi3)
        cv2.imwrite( "fotos/detalhado/roi4"+espaco+".png" , roi4)
        cv2.imwrite( "fotos/detalhado/roi5"+espaco+".png" , roi5)
        cv2.imwrite( "fotos/detalhado/roi6"+espaco+".png" , roi6)"""
    #print(roi4.shape[0]*roi4.shape[1])

    def media_canais(roi):
        # BGR
        azul = 0
        verde = 0
        vermelho = 0
        for a in range(roi.shape[0]):
            for b in range(roi.shape[1]):
                azul = roi[a][b][0] + azul
                verde = roi[a][b][1] + verde
                vermelho = roi[a][b][2] + vermelho

        if(roi.shape[0] == 0 or roi.shape[1]==0):
            return(0,0,0)
        else:
            return (int(azul / (roi.shape[0]*roi.shape[1])), int(verde / (roi.shape[0]*roi.shape[1])), int(vermelho / (roi.shape[0]*roi.shape[1])))

    mroi1 = media_canais(roi1)
    mroi2 = media_canais(roi2)
    mroi3 = media_canais(roi3)
    mroi4 = media_canais(roi4)
    mroi5 = media_canais(roi5)
    mroi6 = media_canais(roi6)

    mediarois = (int((mroi1[0] + mroi2[0] + mroi3[0]) / 3), int((mroi1[1] + mroi2[1] + mroi3[1]) / 3),
                 int((mroi1[2] + mroi2[2] + mroi3[2]) / 3),
                 int((mroi4[0] + mroi5[0] + mroi6[0]) / 3), int((mroi4[1] + mroi5[1] + mroi6[1]) / 3),
                 int((mroi4[2] + mroi5[2] + mroi6[2]) / 3)
                 )
    return (mediarois)

def criar_planilha_excel(ponto,nome,razao,pasta,plan,tempo):
    valrgb = comparaind(nome, ponto, razao, "rgb")
    valhsv = comparaind(nome, ponto, razao, "hsv")
    valhsl = comparaind(nome, ponto, razao, "hls")

    wb = load_workbook(filename=pasta+"/detalhado/resultado.xlsx")
    sh = wb[plan]
    col = 1
    while(sh.cell(row=1, column=col).value!=None):
        col=col+1


    #sh.cell(row=1, column=col).value = nome[6:(len(nome)-4)]+" proximo"
    sh.cell(row=1, column=col).value = plan+ " "+ tempo+" prox"

    sh.cell(row=2, column=col).value = valrgb[2]
    sh.cell(row=3, column=col).value = valrgb[1]
    sh.cell(row=4, column=col).value = valrgb[0]

    sh.cell(row=5, column=col).value = valhsv[0]
    sh.cell(row=6, column=col).value = valhsv[1]
    sh.cell(row=7, column=col).value = valhsv[2]

    sh.cell(row=8, column=col).value = valhsl[0]
    sh.cell(row=9, column=col).value = valhsl[1]
    sh.cell(row=10, column=col).value = valhsl[2]

    #sh.cell(row=1, column=col+1).value = nome[6:(len(nome)-4)]+" distante"
    sh.cell(row=1, column=col+1).value = plan+ " "+ tempo+" dist"

    sh.cell(row=2, column=col+1).value = valrgb[5]
    sh.cell(row=3, column=col+1).value = valrgb[4]
    sh.cell(row=4, column=col+1).value = valrgb[3]

    sh.cell(row=5, column=col+1).value = valhsv[3]
    sh.cell(row=6, column=col+1).value = valhsv[4]
    sh.cell(row=7, column=col+1).value = valhsv[5]

    sh.cell(row=8, column=col+1).value = valhsl[3]
    sh.cell(row=9, column=col+1).value = valhsl[4]
    sh.cell(row=10, column=col+1).value = valhsl[5]

    wb.save(filename=pasta+"/detalhado/resultado.xlsx")


