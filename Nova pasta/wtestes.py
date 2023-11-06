from listar import *
import os
import cv2
from tamanho import *
from excel import *
from selecionar import *

from openpyxl import *



def salvarrois(roi):
    wb = load_workbook(filename="excel.xlsx")
    sh = wb['Sheet1']
    roiredef = cv2.resize(roi, (760, 1100))
    cv2.imshow('img', roiredef)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

    for a in range(roi.shape[0]):#altura - linha
        for b in range(roi.shape[1]):#largura - col

            # 828,320 alturaxlargura
            sh.cell(row=a+1, column=b+1).value = str(roi[a][b])

            roiredef2 = cv2.resize(roi[a][b], (30, 30))


    wb.save(filename= "excel.xlsx")
    """sh.cell(row=1, column=col).value = nome[6:(len(nome) - 4)] + " proximo"

    sh.cell(row=2, column=col).value = valrgb[2]
    sh.cell(row=3, column=col).value = valrgb[1]
    sh.cell(row=4, column=col).value = valrgb[0]

    sh.cell(row=5, column=col).value = valhsv[0]
    sh.cell(row=6, column=col).value = valhsv[1]
    sh.cell(row=7, column=col).value = valhsv[2]

    sh.cell(row=8, column=col).value = valhsl[0]
    sh.cell(row=9, column=col).value = valhsl[1]
    sh.cell(row=10, column=col).value = valhsl[2]

    sh.cell(row=1, column=col + 1).value = nome[6:(len(nome) - 4)] + " distante"

    sh.cell(row=2, column=col + 1).value = valrgb[5]
    sh.cell(row=3, column=col + 1).value = valrgb[4]
    sh.cell(row=4, column=col + 1).value = valrgb[3]

    sh.cell(row=5, column=col + 1).value = valhsv[3]
    sh.cell(row=6, column=col + 1).value = valhsv[4]
    sh.cell(row=7, column=col + 1).value = valhsv[5]

    sh.cell(row=8, column=col + 1).value = valhsl[3]
    sh.cell(row=9, column=col + 1).value = valhsl[4]
    sh.cell(row=10, column=col + 1).value = valhsl[5]

    wb.save(filename=pasta + "/detalhado/resultado.xlsx")"""



def media_canais(roi):
    # BGR
    azul = 0
    verde = 0
    vermelho = 0
    for a in range(roi.shape[0]):
        for b in range(roi.shape[1]):
            azul = roi[a][b][0] + azul
            verde = roi[a][b][1] + verde
            vermelho = roi[a][b][2] + vermelho



    if(roi.shape[0] == 0 or roi.shape[1]==0):
        return(0,0,0)
    else:
        return (int(azul / (roi.shape[0]*roi.shape[1])), int(verde / (roi.shape[0]*roi.shape[1])), int(vermelho / (roi.shape[0]*roi.shape[1])))


"""imgdesenho = cv2.imread('fotos/Sans titre 1.jpg')
cv2.imshow('img', imgdesenho)
cv2.waitKey(0)
cv2.destroyAllWindows()"""

"""tam = imgdesenho.shape#828,320 alturaxlargura
print(tam)"""
"""
salvarrois(cv2.imread('fotos/Sans titre 3.png'))

print("Amarelo - 0 255 255")
print(media_canais(cv2.imread('fotos/amarelo.png')))
print(media_canais(cv2.imread('fotos/amarelo.jpg')))

print("azul - 255 0 0")
print(media_canais(cv2.imread('fotos/azul.png')))
print(media_canais(cv2.imread('fotos/azul.jpg')))

print("ciano - 255 255 0")
print(media_canais(cv2.imread('fotos/ciano.png')))
print(media_canais(cv2.imread('fotos/ciano.jpg')))

print("magenta - 255 0 255")
print(media_canais(cv2.imread('fotos/magenta.png')))
print(media_canais(cv2.imread('fotos/magenta.jpg')))
"""
#[".png", ".bmp", ".dib",  ".jpeg", ".jpg", ".qqjp2",  ".png",  ".webp", ".pbm", ".pgm", ".ppm", ".pxm", ".pnm", ".pfm", ".sr",".ras", ".tiff",".tif" , ".exr ", ".hdr",".pic"]